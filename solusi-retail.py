import pandas as pd
import matplotlib.pyplot as plt
from mlxtend.frequent_patterns import apriori, association_rules
from openpyxl.styles import NamedStyle

# ===============================
# BACA DATA & BUAT DF HARIAN
# ===============================

df = pd.read_excel('data_penjualan.xlsx')
df['tgl_transaksi'] = pd.to_datetime(df['tgl_transaksi'])
df = df.sort_values(['kode_produk', 'tgl_transaksi'])

# DataFrame harian
daily_df = (df.groupby(['tgl_transaksi', 'kode_produk', 'nama_produk'], as_index=False)['total_nilai'].sum())

# Moving average & tren
daily_df['moving_avg'] = (daily_df.groupby('kode_produk')['total_nilai']
                          .transform(lambda x: x.rolling(window=3, min_periods=1).mean()))
daily_df['tren_naik'] = (daily_df['moving_avg'] > daily_df.groupby('kode_produk')['moving_avg'].shift(1))
daily_df['status_tren'] = daily_df['tren_naik'].map({True: 'Naik', False: 'Turun/Tetap'})
daily_df['group_tren'] = (daily_df.groupby('kode_produk')['tren_naik'].transform(lambda x: (~x).cumsum()))
daily_df['jumlah_hari_naik'] = (daily_df.groupby(['kode_produk', 'group_tren']).cumcount() + 1)
daily_df.loc[~daily_df['tren_naik'], 'jumlah_hari_naik'] = 0

# Filter rising: minimal 12 hari naik
daily_df_rising = daily_df[daily_df['jumlah_hari_naik'] >= 12].copy()
daily_df_rising['ma_awal'] = daily_df_rising.groupby('kode_produk')['moving_avg'].transform('first')
daily_df_rising['growth'] = ((daily_df_rising['moving_avg'] / daily_df_rising['ma_awal']) - 1) * 100

df_rising_star = (daily_df_rising.groupby(['kode_produk', 'nama_produk'], as_index=False)
                  .agg({'growth': 'sum', 'total_nilai': 'sum'})
                  .sort_values(by='growth', ascending=False))

rising_star = df_rising_star.copy()
rising_star['Kode Produk'] = df_rising_star['kode_produk']
rising_star['Nama Produk'] = df_rising_star['nama_produk']
rising_star['Growth (%)'] = (
    df_rising_star['growth'].round(2)
)
rising_star['Total Penjualan'] = df_rising_star['total_nilai']

# ===============================
# 2. MARKET BASKET (PACKAGING)
# ===============================

basket = (df.groupby(['nomor_struk', 'nama_produk'])['total_nilai']
          .count()
          .unstack()
          .reset_index()
          .fillna(0)
          .set_index('nomor_struk'))

basket_sets = (basket > 0)

frequent_itemsets = apriori(basket_sets, min_support=0.01, use_colnames=True)
rules = association_rules(frequent_itemsets, metric="lift", min_threshold=1)

daftar_rising_star = set(df_rising_star['nama_produk'].unique())

def cek_rising_star(x):
    return not set(x).isdisjoint(daftar_rising_star)

filtered_rules = rules[
    (rules['lift'] >= 2) &
    (rules['antecedents'].apply(cek_rising_star) | rules['consequents'].apply(cek_rising_star))
].copy()

filtered_rules['antecedents'] = filtered_rules['antecedents'].apply(lambda x: ', '.join(list(x)))
filtered_rules['consequents'] = filtered_rules['consequents'].apply(lambda x: ', '.join(list(x)))

total_struk = len(basket)
filtered_rules['jumlah_invoice'] = (filtered_rules['support'] * total_struk).astype(int)

df_potential_packaging = filtered_rules[
    ['antecedents', 'consequents', 'jumlah_invoice', 'support', 'confidence', 'lift']
]
df_potential_packaging = df_potential_packaging.sort_values(
    by=['lift', 'support', 'confidence'], ascending=[False, False, False]
)

df_potential_packaging = df_potential_packaging.rename(columns={
    'antecedents': 'Jika Membeli',
    'consequents': 'Maka Membeli',
    'jumlah_invoice': 'Jumlah Invoice',
    'support': 'Support',
    'confidence': 'Confidence',
    'lift': 'Lift'
})

# ===============================
# 3. RENAME KOLOM & EXPORT EXCEL
# ===============================
with pd.ExcelWriter('retail_insight.xlsx', engine='openpyxl') as writer:
    rising_star[['Kode Produk', 'Nama Produk', 'Growth (%)', 'Total Penjualan']].to_excel(
        writer, sheet_name='Rising Star', index=False
    )
    workbook = writer.book
    worksheet = writer.sheets['Rising Star']
    style_growth = NamedStyle(
        name="growth_style",
        number_format='0.00'
    )
    workbook.add_named_style(style_growth)
    style_sales = NamedStyle(
        name="sales_style",
        number_format='#,##0'
    )
    workbook.add_named_style(style_sales)
    for row in range(2, worksheet.max_row + 1):
        cell_c = worksheet.cell(row=row, column=3)
        cell_d = worksheet.cell(row=row, column=4)
        cell_c.style = 'growth_style'
        cell_d.style = 'sales_style'

    df_potential_packaging.to_excel(
        writer, sheet_name='Potential Packaging', index=False
    )
    workbook   = writer.book
    worksheet_pp = writer.sheets['Potential Packaging']
    style_metric = NamedStyle(
        name="metric_style",
        number_format='0.00'
    )
    workbook.add_named_style(style_metric)

    for row in range(2, worksheet_pp.max_row + 1):
        for col in range(4, 7):
            cell = worksheet_pp.cell(row=row, column=col)
            cell.style = 'metric_style'

# ============================================================
# VISUALISASI NILAI PERTUMBUHAN RELATIF
# ============================================================

# ----- 1. SPESIFIKASI FIGURE
fig = plt.figure(figsize=(15, 8), dpi=100)
ax = fig.add_subplot(111)

# ----- 2. NORMALIZE BASE 100
def normalize_base100(x):
    base = x.iloc[0] if len(x) > 0 else 1
    return (x / base) * 100

# ----- 3. PALET WARNA
custom_palette = [
    '#FFD700',  # Gold
    '#C0C0C0',  # Silver
    '#CD7F32',  # Bronze
    '#2ecc71',  # Emerald Green
    '#3498db',  # Blue
    '#9b59b6',  # Purple
    '#e74c3c',  # Red
    '#34495e',  # Dark Blue Grey
]
default_color = '#95a5a6'

# ----- 4. PRODUK RISING STAR
df_rising_star_detail = daily_df[
    daily_df['kode_produk'].isin(df_rising_star['kode_produk'])
].copy()

df_rising_star_detail['base100'] = df_rising_star_detail.groupby(
    'kode_produk'
    )['moving_avg'].transform(normalize_base100)

df_rising_star_ranked = df_rising_star.sort_values(by='growth', ascending=False)

color_mapping = {}
rank_mapping = {}
for i, row in enumerate(df_rising_star_ranked.itertuples()):
    kode_produk = row.kode_produk
    color_mapping[kode_produk] = (
        custom_palette[i]
        if i < len(custom_palette)
        else default_color
    )
    rank_mapping[kode_produk] = i + 1

# ----- 5. TOP 3 SALES
top3_sales = (df.groupby(['kode_produk', 'nama_produk'])['total_nilai'].sum().reset_index().sort_values(by='total_nilai', ascending=False).head(3))
top3_codes = top3_sales['kode_produk'].tolist()
top3_plot_df = daily_df[daily_df['kode_produk'].isin(top3_codes)].copy()

top3_plot_df['base100'] = top3_plot_df.groupby('kode_produk')['moving_avg'].transform(normalize_base100)

# ----- 6. PLOT TOP 3 SALES
grey_colors = [
    '#B0B0B0',
    '#909090',
    '#707070'
]

for idx, (kode_produk, group) in enumerate(
    top3_plot_df.groupby('kode_produk')
):
    nama_produk = group['nama_produk'].iloc[0]
    grey_color = (
        grey_colors[idx]
        if idx < len(grey_colors)
        else '#808080'
    )
    ax.plot(
        group['tgl_transaksi'],
        group['base100'],
        linestyle='--',
        linewidth=2,
        marker='o',
        markersize=3,
        color=grey_color,
        alpha=0.7,
        label=f"Top Sales: {nama_produk}"
    )

# ----- 7. PLOT RISING STAR
for kode_produk, group in df_rising_star_detail.groupby('kode_produk'):
    nama_produk = group['nama_produk'].iloc[0]
    line_color = color_mapping.get(
        kode_produk, default_color
    )
    rank = rank_mapping.get(
        kode_produk,
        '?'
    )
    label_with_rank = f'Rank {rank}: {nama_produk}'
    ax.plot(
        group['tgl_transaksi'],
        group['base100'],
        marker='o',
        markersize=4,
        linewidth=2.5,
        color=line_color,
        label=label_with_rank
    )

# ----- 8. JUDUL DAN LABEL GRAFIK
font_title = {
    'family': 'sans-serif',
    'color': 'black',
    'weight': 'bold',
    'size': 16
}
font_label = {
    'family': 'sans-serif',
    'weight': 'normal',
    'size': 12
}
ax.set_title(
    'ANALISIS PERTUMBUHAN RELATIF PRODUK RISING STAR\n'
    '(Dengan Benchmark Top 3 Total Penjualan)',
    fontdict=font_title,
    pad=20
)
ax.set_xlabel(
    'Periode Tanggal',
    fontdict=font_label,
    labelpad=10
)
ax.set_ylabel(
    'Indeks Pertumbuhan (Base 100)',
    fontdict=font_label,
    labelpad=10
)

# ----- 9. GRID & BASELINE

ax.grid(
    True,
    linestyle='--',
    linewidth=0.5,
    alpha=0.5
)
ax.axhline(
    y=100,
    color='black',
    linestyle='-',
    linewidth=1,
    alpha=0.5
)

# ------ 10. FORMAT AXIS

plt.xticks(
    rotation=45,
    ha='right',
    fontsize=10
)
plt.yticks(fontsize=10)

# ----- 11. SORT LEGEND BERDASARKAN RANK

handles, labels = ax.get_legend_handles_labels()
top_sales_items = []
rising_items = []

for h, l in zip(handles, labels):
    if l.startswith('Top Sales'):
        top_sales_items.append((h, l))
    else:
        rising_items.append((h, l))

rising_items = sorted(
    rising_items,
    key=lambda x: int(
        x[1].split(':')[0].split()[1]
    )
)

final_legend = top_sales_items + rising_items

final_handles = [x[0] for x in final_legend]
final_labels = [x[1] for x in final_legend]

# ----- 12. LEGEND

ax.legend(
    final_handles,
    final_labels,
    title="Kategori Produk",
    title_fontsize=12,
    fontsize=10,
    bbox_to_anchor=(1.02, 1),
    loc='upper left',
    borderaxespad=0,
    frameon=True,
    shadow=True
)

# ----- 13. LAYOUT & SAVE

plt.tight_layout()
plt.savefig(
    'rising_star_index.png',
    bbox_inches='tight'
)

# ===============================
# VISUALISASI NILAI PENJUALAN ASLI
# ===============================

fig2 = plt.figure(figsize=(15, 8), dpi=100)
ax2 = fig2.add_subplot(111)

# ----- 1. PLOT TOP 3 SALES

for idx2, (kode_produk, group) in enumerate(
    top3_plot_df.groupby('kode_produk')
):
    nama_produk = group['nama_produk'].iloc[0]
    grey_color = (
        grey_colors[idx2]
        if idx2 < len(grey_colors)
        else '#808080'
    )
    ax2.plot(
        group['tgl_transaksi'],
        group['total_nilai'],
        linestyle='--',
        linewidth=2,
        marker='o',
        markersize=3,
        color=grey_color,
        alpha=0.7,
        label=f"Top Sales: {nama_produk}"
    )

# ----- 2. PLOT RISING STAR
for kode_produk, group in df_rising_star_detail.groupby('kode_produk'):
    nama_produk = group['nama_produk'].iloc[0]
    line_color = color_mapping.get(
        kode_produk,
        default_color
    )
    rank = rank_mapping.get(
        kode_produk,
        '?'
    )
    label_with_rank = f'Rank {rank}: {nama_produk}'
    ax2.plot(
        group['tgl_transaksi'],
        group['total_nilai'],
        marker='o',
        markersize=4,
        linewidth=2.5,
        color=line_color,
        label=label_with_rank
    )

# ----- 3. JUDUL DAN LABEL GRAFIK
ax2.set_title(
    'ANALISIS NILAI PENJUALAN PRODUK RISING STAR\n'
    '(Nilai Penjualan Asli)',
    fontdict=font_title,
    pad=20
)
ax2.set_xlabel(
    'Periode Tanggal',
    fontdict=font_label,
    labelpad=10
)
ax2.set_ylabel(
    'Total Nilai Penjualan',
    fontdict=font_label,
    labelpad=10
)

# ----- 4. GRID
ax2.grid(
    True,
    linestyle='--',
    linewidth=0.5,
    alpha=0.5
)

# ----- 5. FORMAT AXIS
plt.xticks(
    rotation=45,
    ha='right',
    fontsize=10
)
plt.yticks(fontsize=10)

# ----- 6. SORT LEGEND
handles2, labels2 = ax2.get_legend_handles_labels()
top_sales_items2 = []
rising_items2 = []

for h, l in zip(handles2, labels2):
    if l.startswith('Top Sales'):
        top_sales_items2.append((h, l))
    else:
        rising_items2.append((h, l))

rising_items2 = sorted(
    rising_items2,
    key=lambda x: int(
        x[1].split(':')[0].split()[1]
    )
)

final_legend2 = top_sales_items2 + rising_items2

final_handles2 = [x[0] for x in final_legend2]
final_labels2 = [x[1] for x in final_legend2]

# ----- 7. LEGEND
ax2.legend(
    final_handles2,
    final_labels2,
    title="Kategori Produk",
    title_fontsize=12,
    fontsize=10,
    bbox_to_anchor=(1.02, 1),
    loc='upper left',
    borderaxespad=0,
    frameon=True,
    shadow=True
)

# ----- 8. LAYOUT & SAVE
plt.tight_layout()
plt.savefig(
    'rising_star_actual.png',
    bbox_inches='tight'
)